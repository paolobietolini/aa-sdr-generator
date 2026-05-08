from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from exporters.excel import (
    COL_DESCRIPTION,
    COL_NAME,
    COL_VARIABLE,
    DATA_START_ROW,
    _cm_format,
    _fill_sheet,
    _index_by_short_id,
    _output_filename,
    _segment_container,
    _short_id,
    filter_suites,
)
from models.adobe.analytics import (
    CalculatedMetricResponse,
    DimensionResponse,
    SegmentDefinition,
    SegmentResponse,
    SuiteResponse,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _suite(rsid: str) -> SuiteResponse:
    return SuiteResponse(rsid=rsid, id=rsid, name=rsid)


def _ws(rows: list[tuple]) -> openpyxl.worksheet.worksheet.Worksheet:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row, col, val in rows:
        ws.cell(row, col).value = val
    return ws


# ---------------------------------------------------------------------------
# filter_suites
# ---------------------------------------------------------------------------

def test_filter_suites_include_glob():
    suites = [_suite("myco_be"), _suite("myco_de"), _suite("other")]
    result = filter_suites(suites, include=["myco*"], exclude=[])
    assert [s.rsid for s in result] == ["myco_be", "myco_de"]


def test_filter_suites_exclude_glob():
    suites = [_suite("myco_be"), _suite("myco_dev"), _suite("myco_de")]
    result = filter_suites(suites, include=["myco*"], exclude=["*_dev"])
    assert [s.rsid for s in result] == ["myco_be", "myco_de"]


def test_filter_suites_no_match_returns_empty():
    suites = [_suite("other1"), _suite("other2")]
    result = filter_suites(suites, include=["myco*"], exclude=[])
    assert result == []


def test_filter_suites_wildcard_include():
    suites = [_suite("anything"), _suite("else")]
    result = filter_suites(suites, include=["*"], exclude=[])
    assert len(result) == 2


def test_filter_suites_empty_input():
    assert filter_suites([], include=["*"], exclude=[]) == []


def test_filter_suites_exclude_all():
    suites = [_suite("myco_be")]
    result = filter_suites(suites, include=["*"], exclude=["*"])
    assert result == []


def test_filter_suites_multiple_include_patterns():
    suites = [_suite("be_prod"), _suite("de_prod"), _suite("other")]
    result = filter_suites(suites, include=["be_*", "de_*"], exclude=[])
    assert [s.rsid for s in result] == ["be_prod", "de_prod"]


# ---------------------------------------------------------------------------
# _short_id / _index_by_short_id
# ---------------------------------------------------------------------------

def test_short_id_strips_prefix():
    assert _short_id("variables/evar1") == "evar1"


def test_short_id_no_prefix_unchanged():
    assert _short_id("evar1") == "evar1"


def test_index_by_short_id_strips_prefix_and_casefolds():
    dim = DimensionResponse(id="variables/eVar1", title="Page", name="page", description="desc")
    result = _index_by_short_id([dim])
    assert "evar1" in result
    assert result["evar1"] is dim


def test_index_by_short_id_multiple_items():
    dims = [
        DimensionResponse(id="variables/eVar1", title="A", name="a", description=""),
        DimensionResponse(id="variables/prop3", title="B", name="b", description=""),
    ]
    result = _index_by_short_id(dims)
    assert set(result) == {"evar1", "prop3"}


# ---------------------------------------------------------------------------
# _fill_sheet
# ---------------------------------------------------------------------------

def test_fill_sheet_writes_name_and_description():
    ws = _ws([(DATA_START_ROW, COL_VARIABLE, "eVar1")])
    item = MagicMock(title="Page Name", name="page", description="The page name")
    filled = _fill_sheet(ws, {"evar1": item})
    assert filled == 1
    assert ws.cell(DATA_START_ROW, COL_NAME).value == "Page Name"
    assert ws.cell(DATA_START_ROW, COL_DESCRIPTION).value == "The page name"


def test_fill_sheet_falls_back_to_name_when_title_none():
    ws = _ws([(DATA_START_ROW, COL_VARIABLE, "eVar1")])
    item = MagicMock()
    item.title = None
    item.name = "fallback"
    item.description = "desc"
    _fill_sheet(ws, {"evar1": item})
    assert ws.cell(DATA_START_ROW, COL_NAME).value == "fallback"


def test_fill_sheet_skips_unmatched_row():
    ws = _ws([
        (DATA_START_ROW, COL_VARIABLE, "eVar1"),
        (DATA_START_ROW + 1, COL_VARIABLE, "eVar2"),
    ])
    item = MagicMock(title="Name", name="name", description="desc")
    filled = _fill_sheet(ws, {"evar1": item})
    assert filled == 1
    assert ws.cell(DATA_START_ROW + 1, COL_NAME).value is None


def test_fill_sheet_skips_empty_variable_cell():
    ws = _ws([
        (DATA_START_ROW, COL_VARIABLE, None),
        (DATA_START_ROW + 1, COL_VARIABLE, "eVar1"),
    ])
    item = MagicMock(title="Name", name="name", description="desc")
    filled = _fill_sheet(ws, {"evar1": item})
    assert filled == 1


def test_fill_sheet_alias_maps_variable_id():
    ws = _ws([(DATA_START_ROW, COL_VARIABLE, "pageName")])
    item = MagicMock(title="Page", name="page", description="The page")
    filled = _fill_sheet(ws, {"page": item}, aliases={"pagename": "page"})
    assert filled == 1
    assert ws.cell(DATA_START_ROW, COL_NAME).value == "Page"


def test_fill_sheet_is_case_insensitive():
    ws = _ws([(DATA_START_ROW, COL_VARIABLE, "EVAR1")])
    item = MagicMock(title="Name", name="name", description="desc")
    filled = _fill_sheet(ws, {"evar1": item})
    assert filled == 1


def test_fill_sheet_returns_zero_for_no_matches():
    ws = _ws([(DATA_START_ROW, COL_VARIABLE, "eVar99")])
    filled = _fill_sheet(ws, {})
    assert filled == 0


# ---------------------------------------------------------------------------
# _cm_format
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("raw,expected", [
    ("decimal", "Decimal"),
    ("percent", "Percent"),
    ("time", "Time"),
    ("currency", "Currency"),
])
def test_cm_format_known_types(raw, expected):
    cm = CalculatedMetricResponse(id="cm1", type=raw)
    assert _cm_format(cm) == expected


def test_cm_format_unknown_type_returns_raw():
    cm = CalculatedMetricResponse(id="cm1", type="custom_type")
    assert _cm_format(cm) == "custom_type"


def test_cm_format_none_type_returns_none():
    cm = CalculatedMetricResponse(id="cm1", type=None)
    assert _cm_format(cm) is None


# ---------------------------------------------------------------------------
# _segment_container
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("context,expected", [
    ("hits", "Hit"),
    ("visits", "Visit"),
    ("visitors", "Visitor"),
])
def test_segment_container_known_contexts(context, expected):
    seg = SegmentResponse(id="s1", definition=SegmentDefinition(container={"context": context}))
    assert _segment_container(seg) == expected


def test_segment_container_unknown_context_returns_none():
    seg = SegmentResponse(id="s1", definition=SegmentDefinition(container={"context": "unknown"}))
    assert _segment_container(seg) is None


def test_segment_container_no_definition_returns_none():
    seg = SegmentResponse(id="s1", definition=None)
    assert _segment_container(seg) is None


def test_segment_container_no_container_returns_none():
    seg = SegmentResponse(id="s1", definition=SegmentDefinition(container=None))
    assert _segment_container(seg) is None


# ---------------------------------------------------------------------------
# _output_filename
# ---------------------------------------------------------------------------

def test_output_filename_without_author():
    with patch("exporters.excel._date") as mock_date:
        mock_date.today.return_value.strftime.return_value = "2026-05-08"
        name = _output_filename("myco_be", None)
    assert name == "myco_be_2026-05-08_sdr.xlsx"


def test_output_filename_with_author():
    with patch("exporters.excel._date") as mock_date:
        mock_date.today.return_value.strftime.return_value = "2026-05-08"
        name = _output_filename("myco_be", "Paolo")
    assert name == "myco_be_2026-05-08_Paolo_sdr.xlsx"


def test_output_filename_author_sanitised():
    with patch("exporters.excel._date") as mock_date:
        mock_date.today.return_value.strftime.return_value = "2026-05-08"
        name = _output_filename("myco_be", "John Doe")
    assert name == "myco_be_2026-05-08_John_Doe_sdr.xlsx"
