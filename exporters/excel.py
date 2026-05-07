from fnmatch import fnmatch
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.workbook import Workbook

from core.client import AdobeAnalyticsClient
from models.adobe.analytics import CalculatedMetricResponse, SegmentResponse, SuiteResponse
from config.sdr_config import SdrConfig


# eVars, props, and custom events all share this layout:
# col 3 = identifier, col 4 = name, col 5 = description, data starts row 7
COL_VARIABLE = 3
COL_NAME = 4
COL_DESCRIPTION = 5
DATA_START_ROW = 7

PROP_ALIASES = {
    "pagename": "page",
}

# metrics-segments sheet columns
MS_COL_TYPE = 3
MS_COL_NAME = 4
MS_COL_DESCRIPTION = 5
MS_COL_FORMAT = 6

CONTAINER_LEVEL_MAP = {"hits": "Hit", "visits": "Visit", "visitors": "Visitor"}
CM_FORMAT_MAP = {"decimal": "Decimal", "percent": "Percent", "time": "Time", "currency": "Currency"}

GLOSSARY_ORG_CELL = (2, 3)


def filter_suites(suites: Iterable[SuiteResponse], include: list[str], exclude: list[str]) -> list[SuiteResponse]:
    matched = [s for s in suites if any(fnmatch(s.rsid, p) for p in include)]
    if exclude:
        matched = [s for s in matched if not any(fnmatch(s.rsid, p) for p in exclude)]
    return matched


def _short_id(full_id: str) -> str:
    return full_id.split("/")[-1] if "/" in full_id else full_id


def _index_by_short_id(items: list[Any]) -> dict[str, Any]:
    return {_short_id(item.id).casefold(): item for item in items}


def _fill_sheet(ws, by_sid: dict[str, Any], aliases: dict[str, str] | None = None) -> int:
    aliases = aliases or {}
    filled = 0
    for row in range(DATA_START_ROW, ws.max_row + 1):
        var_id = ws.cell(row, COL_VARIABLE).value
        if not var_id:
            continue
        key = str(var_id).casefold()
        key = aliases.get(key, key)
        match = by_sid.get(key)
        if not match:
            continue
        ws.cell(row, COL_NAME).value = match.title or match.name
        ws.cell(row, COL_DESCRIPTION).value = match.description
        filled += 1
    return filled


def _cm_format(cm: CalculatedMetricResponse) -> str | None:
    return CM_FORMAT_MAP.get((cm.type or "").lower(), cm.type)


def _segment_container(seg: SegmentResponse) -> str | None:
    if not seg.definition or not seg.definition.container:
        return None
    return CONTAINER_LEVEL_MAP.get(seg.definition.container.get("context"))


def _fill_metrics_segments_sheet(
    ws,
    calc_metrics: list[CalculatedMetricResponse],
    segments: list[SegmentResponse],
) -> tuple[int, int]:
    row = DATA_START_ROW

    for cm in calc_metrics:
        ws.cell(row, MS_COL_TYPE).value = "Calculated Metric"
        ws.cell(row, MS_COL_NAME).value = cm.name
        ws.cell(row, MS_COL_DESCRIPTION).value = cm.description
        ws.cell(row, MS_COL_FORMAT).value = _cm_format(cm)
        row += 1

    for seg in segments:
        ws.cell(row, MS_COL_TYPE).value = "Segment"
        ws.cell(row, MS_COL_NAME).value = seg.name
        ws.cell(row, MS_COL_DESCRIPTION).value = seg.description
        ws.cell(row, MS_COL_FORMAT).value = _segment_container(seg)
        row += 1

    # Clear leftover sample rows from template
    for r in range(row, ws.max_row + 1):
        for c in [MS_COL_TYPE, MS_COL_NAME, MS_COL_DESCRIPTION, MS_COL_FORMAT]:
            ws.cell(r, c).value = None

    return len(calc_metrics), len(segments)


def _set_org_name(wb: Workbook, name: str) -> None:
    wb["Glossary"].cell(*GLOSSARY_ORG_CELL).value = name
    wb["reserved reporting"].cell(*GLOSSARY_ORG_CELL).value = name


def _resolve_org_name(client: AdobeAnalyticsClient, override: str | None) -> str:
    if override:
        return override
    me = client.client.discover_me()
    return me.ims_orgs[0].companies[0].company_name


def generate_sdr(client: AdobeAnalyticsClient, config: SdrConfig) -> list[Path]:
    suites = client.get_suites()
    selected = filter_suites(suites, config.rsids.include, config.rsids.exclude)
    if not selected:
        print(f"No report suites matched include={config.rsids.include} exclude={config.rsids.exclude}")
        return []

    org_name = _resolve_org_name(client, config.metadata.organization)
    config.output_dir.mkdir(parents=True, exist_ok=True)

    output_files: list[Path] = []
    for suite in selected:
        wb = openpyxl.load_workbook(config.template_path)
        _set_org_name(wb, org_name)

        dims = client.get_dimensions(rsid=suite.rsid, expansion="description,tags")
        dims_by_sid = _index_by_short_id(dims)
        n_evars = _fill_sheet(wb["eVars"], dims_by_sid)
        n_props = _fill_sheet(wb["props"], dims_by_sid, aliases=PROP_ALIASES)

        metrics = client.get_metrics(rsid=suite.rsid, expansion="description")
        metrics_by_sid = _index_by_short_id(metrics)
        n_events = _fill_sheet(wb["custom events (metrics)"], metrics_by_sid)

        cms = client.get_calculated_metrics(rsids=suite.rsid, expansion="definition,reportSuiteName")
        segs = client.get_segments(rsids=suite.rsid, expansion="definition,reportSuiteName")
        n_cms, n_segs = _fill_metrics_segments_sheet(wb["metrics-segments"], cms, segs)

        out_path = config.output_dir / f"{suite.rsid}_sdr.xlsx"
        wb.save(out_path)
        output_files.append(out_path)
        print(f"  {suite.rsid}: eVars={n_evars} props={n_props} events={n_events} cms={n_cms} segs={n_segs} -> {out_path}")

    return output_files
